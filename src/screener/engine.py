from __future__ import annotations

import logging
import os
import sqlite3
from pathlib import Path
from typing import Callable

import numpy as np
import pandas as pd
import yaml
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parents[2]
CONFIG_PATH = BASE_DIR / "config" / "screener_config.yaml"
DB_PATH = Path(os.environ.get("DB_PATH", BASE_DIR / "data" / "nifty100.db"))
OUTPUT_DIR = BASE_DIR / "output"
EXCEL_PATH = OUTPUT_DIR / "screener_output.xlsx"

FINANCIALS_SECTOR = "Financials"
DEBT_FREE_LABEL = "Debt Free"
DEFAULT_RANK_BY = "composite_quality_score"
PRESET_COUNT_BAND = (5, 50)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

FILTER_SPECS: dict[str, tuple[str, str]] = {
    "roe_min": ("return_on_equity_pct", "min"),
    "de_max": ("debt_to_equity", "max"),
    "fcf_min": ("free_cash_flow_cr", "min"),
    "revenue_cagr_3yr_min": ("revenue_cagr_3yr", "min"),
    "revenue_cagr_5yr_min": ("revenue_cagr_5yr", "min"),
    "pat_cagr_5yr_min": ("pat_cagr_5yr", "min"),
    "opm_min": ("operating_profit_margin_pct", "min"),
    "pe_max": ("pe_ratio", "max"),
    "pb_max": ("pb_ratio", "max"),
    "dividend_yield_min": ("dividend_yield_pct", "min"),
    "dividend_payout_max": ("dividend_payout_ratio_pct", "max"),
    "icr_min": ("interest_coverage", "min"),
    "market_cap_min": ("market_cap_crore", "min"),
    "net_profit_min": ("net_profit", "min"),
    "eps_cagr_min": ("eps_cagr_5yr", "min"),
    "asset_turnover_min": ("asset_turnover", "min"),
    "sales_min": ("sales", "min"),
    "de_declining": ("de_declining", "bool"),
}

WINSORISED_METRICS: dict[str, float] = {
    "return_on_equity_pct": 0.15,
    "return_on_capital_pct": 0.10,
    "net_profit_margin_pct": 0.10,
    "fcf_cagr_5yr": 0.15,
    "cfo_to_pat": 0.10,
    "revenue_cagr_5yr": 0.10,
    "pat_cagr_5yr": 0.10,
}

ABSOLUTE_METRICS: list[tuple[str, float, Callable[[pd.Series], pd.Series]]] = []

EXPORT_COLUMNS = [
    "company_id", "company_name", "sector", "year",
    "return_on_equity_pct", "return_on_capital_pct", "net_profit_margin_pct",
    "operating_profit_margin_pct", "debt_to_equity", "interest_coverage",
    "asset_turnover", "free_cash_flow_cr", "fcf_cagr_5yr", "cfo_to_pat",
    "revenue_cagr_3yr", "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "pe_ratio", "pb_ratio", "dividend_yield_pct", "dividend_payout_ratio_pct",
    "market_cap_crore", "sales", "net_profit", "composite_quality_score",
]

RATIOS_QUERY = """
    SELECT r.*, s.broad_sector AS sector, c.company_name,
           m.market_cap_crore, m.pe_ratio, m.pb_ratio, m.dividend_yield_pct,
           p.net_profit, p.sales, p.operating_profit, p.depreciation, p.eps,
           b.equity_capital, b.reserves, b.borrowings
    FROM financial_ratios r
    LEFT JOIN sectors s ON s.company_id = r.company_id
    LEFT JOIN companies c ON c.id = r.company_id
    LEFT JOIN market_cap m
        ON m.company_id = r.company_id
       AND m.year = CAST(SUBSTR(r.year, 1, 4) AS INTEGER)
    LEFT JOIN profitandloss p USING (company_id, year)
    LEFT JOIN balancesheet b USING (company_id, year)
"""


def load_config(path: str | Path = CONFIG_PATH) -> dict:
    """Load screener thresholds from YAML."""
    with open(path, encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def load_financial_ratios(db_path: str | Path = DB_PATH) -> pd.DataFrame:
    """Load all-year ratios joined with sector, valuation, P&L and BS fields."""
    with sqlite3.connect(db_path) as conn:
        return pd.read_sql_query(RATIOS_QUERY, conn)


def resolve_preset(config: dict) -> tuple[str, dict, str]:
    """Return (name, filters, rank_by) for the active preset or custom filters."""
    presets = config.get("presets") or {}
    active = config.get("active_preset")
    if active not in presets:
        return "custom", config.get("filters") or {}, DEFAULT_RANK_BY
    preset = presets[active] or {}
    if "filters" in preset or "rank_by" in preset:
        return active, preset.get("filters") or {}, preset.get("rank_by", DEFAULT_RANK_BY)
    return active, preset, DEFAULT_RANK_BY


def _cagr_series(frame: pd.DataFrame, column: str, years: int) -> pd.Series:
    """CAGR with doc edge rules: None unless base > 0 and end > 0."""
    ordered = frame.sort_values(["company_id", "year"])
    base = ordered.groupby("company_id")[column].shift(years)
    end = ordered[column]
    cagr = ((end / base) ** (1 / years) - 1) * 100
    return cagr.where((base > 0) & (end > 0))


def _add_derived_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Compute score inputs missing from the ratios table (multi-year aware)."""
    result = df.copy()
    num = lambda col: pd.to_numeric(result[col], errors="coerce") if col in result.columns else None
    if "return_on_capital_pct" not in result.columns and all(
        c in result.columns for c in ("operating_profit", "depreciation", "equity_capital", "reserves", "borrowings")
    ):
        ebit = num("operating_profit") - num("depreciation")
        capital = num("equity_capital") + num("reserves") + num("borrowings")
        result["return_on_capital_pct"] = (ebit / capital.replace(0, np.nan) * 100).where(capital > 0)
    if "cfo_to_pat" not in result.columns and all(
        c in result.columns for c in ("cash_from_operations_cr", "net_profit")
    ):
        pat = num("net_profit")
        result["cfo_to_pat"] = (num("cash_from_operations_cr") / pat.replace(0, np.nan)).where(pat != 0)
    if "fcf_cagr_5yr" not in result.columns and "free_cash_flow_cr" in result.columns:
        result["fcf_cagr_5yr"] = _cagr_series(result, "free_cash_flow_cr", 5)
    for column, source, years in [
        ("revenue_cagr_3yr", "sales", 3),
        ("revenue_cagr_5yr", "sales", 5),
        ("pat_cagr_5yr", "net_profit", 5),
        ("eps_cagr_5yr", "eps", 5),
    ]:
        if column not in result.columns and source in result.columns:
            result[column] = _cagr_series(result, source, years)
    if {"free_cash_flow_cr", "market_cap_crore"} <= set(result.columns):
        fcf_yield = num("free_cash_flow_cr") / num("market_cap_crore") * 100
        result["fcf_yield_pct"] = fcf_yield.replace([np.inf, -np.inf], np.nan).round(2)
    return result


def latest_year_frame(df: pd.DataFrame) -> pd.DataFrame:
    """Flag declining D/E YoY and keep the latest year per company."""
    frame = df.sort_values(["company_id", "year"]).copy()
    prior_de = frame.groupby("company_id", group_keys=False)["debt_to_equity"].shift(1)
    frame["de_declining"] = pd.to_numeric(frame["debt_to_equity"], errors="coerce").lt(prior_de)
    return frame.groupby("company_id", group_keys=False).tail(1)


def _coerce(df: pd.DataFrame, column: str) -> pd.Series:
    series = df[column]
    if column == "interest_coverage":
        series = pd.to_numeric(series.replace(DEBT_FREE_LABEL, np.inf), errors="coerce")
        return series.fillna(np.inf)
    return pd.to_numeric(series, errors="coerce")


def apply_filters(df: pd.DataFrame, filters: dict | None) -> pd.DataFrame:
    """Apply YAML threshold filters to the ratios DataFrame."""
    mask = pd.Series(True, index=df.index)
    for key, value in (filters or {}).items():
        if value is None or key not in FILTER_SPECS:
            continue
        column, direction = FILTER_SPECS[key]
        if column not in df.columns:
            logger.warning("Filter %s skipped: missing column %s", key, column)
            continue
        if direction == "bool":
            if value is not True:
                continue
            mask &= df[column].fillna(False).astype(bool)
        else:
            if key == "de_max" and "sector" in df.columns:
                mask &= df["sector"].ne(FINANCIALS_SECTOR)
            series = _coerce(df, column)
            mask &= series.ge(value) if direction == "min" else series.le(value)
        logger.info("%-24s -> %d companies", key, int(mask.sum()))
    return df.loc[mask]


def _winsorised_score(series: pd.Series) -> pd.Series:
    """P10/P90 winsorise then scale to 0-100."""
    lo, hi = series.quantile((0.1, 0.9))
    if not (np.isfinite(lo) and np.isfinite(hi)) or hi <= lo:
        return pd.Series(50.0, index=series.index)
    return ((series.clip(lo, hi) - lo) / (hi - lo) * 100).fillna(0.0)


def _fcf_flag_score(series: pd.Series) -> pd.Series:
    """100 if FCF positive else 0."""
    return series.gt(0).astype(float) * 100.0


def _de_score(series: pd.Series) -> pd.Series:
    """Piecewise D/E score: 0=100, 0.5=85, 1=70, 2=50, >5=0."""
    scored = np.interp(series.clip(0, 5).to_numpy(), [0, 0.5, 1, 2, 5], [100, 85, 70, 50, 0])
    return pd.Series(scored, index=series.index).fillna(0.0)


def _icr_score(series: pd.Series) -> pd.Series:
    """Piecewise ICR score: <1.5=0, 3=50, 5=75, >10=100 (Debt Free=100)."""
    scored = np.interp(series.clip(1.5, 10).to_numpy(), [1.5, 3, 5, 10], [0, 50, 75, 100])
    return pd.Series(scored, index=series.index).fillna(0.0)


ABSOLUTE_METRICS.extend([
    ("free_cash_flow_cr", 0.05, _fcf_flag_score),
    ("debt_to_equity", 0.10, _de_score),
    ("interest_coverage", 0.05, _icr_score),
])


def add_composite_score(df: pd.DataFrame) -> pd.DataFrame:
    """Add 0-100 composite score per doc 25.1, sector-relative winsorisation."""
    result = df.copy()
    sector = result["sector"].fillna("Unassigned") if "sector" in result.columns else None
    score = pd.Series(0.0, index=result.index)
    total = 0.0
    for column, weight in WINSORISED_METRICS.items():
        if column not in result.columns:
            continue
        series = _coerce(result, column)
        scored = series.groupby(sector).transform(_winsorised_score) if sector is not None else _winsorised_score(series)
        score += weight * scored
        total += weight
    for column, weight, func in ABSOLUTE_METRICS:
        if column not in result.columns:
            continue
        score += weight * func(_coerce(result, column))
        total += weight
    result["composite_quality_score"] = (score / total).round(2) if total else np.nan
    return result


def screen(financial_ratios: pd.DataFrame, config: dict | None = None) -> pd.DataFrame:
    """Score the full universe, filter per active preset, rank and return."""
    config = config if config is not None else load_config()
    _, filters, rank_by = resolve_preset(config)
    frame = latest_year_frame(_add_derived_metrics(financial_ratios))
    frame = apply_filters(add_composite_score(frame), filters)
    if rank_by not in frame.columns:
        rank_by = DEFAULT_RANK_BY
    return frame.sort_values(rank_by, ascending=False, na_position="last").reset_index(drop=True)


def run_all_presets(df: pd.DataFrame, config: dict) -> dict[str, pd.DataFrame]:
    """Run every preset and validate result counts against the 5-50 band."""
    results = {}
    for name in (config.get("presets") or {}):
        result = screen(df, {**config, "active_preset": name})
        lo, hi = PRESET_COUNT_BAND
        status = "OK" if lo <= len(result) <= hi else "OUT OF BAND"
        logger.info("Preset %-20s -> %2d companies [%s]", name, len(result), status)
        results[name] = result
    return results


def _style_sheet(ws, frame: pd.DataFrame, filters: dict) -> None:
    """Green fill for cells meeting preset thresholds, red for failing."""
    thresholds = {
        FILTER_SPECS[k][0]: (k, v)
        for k, v in (filters or {}).items()
        if k in FILTER_SPECS and FILTER_SPECS[k][1] != "bool" and v is not None
    }
    for col_idx, column in enumerate(frame.columns, start=1):
        if column not in thresholds:
            continue
        key, value = thresholds[column]
        direction = FILTER_SPECS[key][1]
        for row_idx, raw in enumerate(frame[column].tolist(), start=2):
            value_ok = pd.notna(raw) and (raw >= value if direction == "min" else raw <= value)
            ws.cell(row=row_idx, column=col_idx).fill = GREEN_FILL if value_ok else RED_FILL


def main() -> None:
    """Run all presets, export colour-coded screener_output.xlsx (D-07)."""
    logging.basicConfig(level=logging.INFO)
    config = load_config()
    ratios = load_financial_ratios()
    results = run_all_presets(ratios, config)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(EXCEL_PATH) as writer:
        for name, frame in results.items():
            _, filters, _ = resolve_preset({**config, "active_preset": name})
            export = frame.sort_values("composite_quality_score", ascending=False)
            export = export[[c for c in EXPORT_COLUMNS if c in export.columns]]
            sheet = name[:31]
            export.to_excel(writer, sheet_name=sheet, index=False)
            _style_sheet(writer.sheets[sheet], export, filters)
    logger.info("Wrote %s", EXCEL_PATH)
    print(results[config.get("active_preset", "")][["company_id", "sector", "composite_quality_score"]].head(10).to_string(index=False))


if __name__ == "__main__":
    main()